
# coding: UTF-8

# extract_data_from_files_pdf's.py

import os
from openpyxl import Workbook
import pdfplumber
import re
from datetime import datetime
import mysql.connector

def execute_insert (cursor, invoice_number, invoice_date, file_name, status):
    
    sql = 'INSERT INTO invoices (invoice_number, invoice_date, file_name, status) VALUES (%s, %s, %s, %s);'
    prepare = (invoice_number, invoice_date, file_name, status)
    
    cursor.execute (sql, prepare)

def main ():

    database = mysql.connector.connect (
        host = 'localhost',
        user = 'root',
        password = '',
        database = 'invoices'
    )
    
    cursor = database.cursor ()

    directory = "invoices_pdf's"
    files = os.listdir (directory)
    files_quantity = len (files)

    if files_quantity == 0:
        
        raise Exception ('No files found in the directory!')

    wb = Workbook ()
    ws = wb.active
    ws.title = "PDF's invoices"

    ws['A1'] = 'INVOICE'
    ws['B1'] = 'DATE'
    ws['C1'] = 'FILE'
    ws['D1'] = 'STATUS'

    last_empty_line = 1
    
    while ws['D' + str(last_empty_line)].value is not None:
        
        last_empty_line += 1

    for file in files:
        
        try:
        
            with pdfplumber.open (directory + '/' + file) as pdf:
                
                first_page = pdf.pages[0]
                pdf_text = first_page.extract_text ()

            invoice_number_re_pattern = r'INVOICE #(\d+)'
            invoice_date_re_pattern = r'DATE: (\d{2}/\d{2}/\d{4})'

            match_number = re.search (invoice_number_re_pattern, pdf_text)
            match_date = re.search (invoice_date_re_pattern, pdf_text)

            if match_number:
                
                ws['A{}'.format (last_empty_line)] = match_number.group (1)
                
            else:
                
                raise Exception ("Couldn't find invoice number!")

            if match_date:
                
                ws['B{}'.format (last_empty_line)] = match_date.group (1)
                
            else:
                
                raise Exception ("Couldn't find invoice date!")

            ws['C{}'.format (last_empty_line)] = file
            ws['D{}'.format (last_empty_line)] = 'completed!'

            execute_insert (cursor, match_number.group (1), match_date.group (1), file, 'completed!')
            
            database.commit ()

            last_empty_line += 1

        except Exception as error:
            
            print(f'Error processing file: {error}')

            ws['C{}'.format (last_empty_line)] = file
            ws['D{}'.format (last_empty_line)] = 'Exception: {}'.format (error)

            execute_insert (cursor, 'N/A', 'N/A', file, 'Exception: {}'.format (error))
            
            database.commit ()

            last_empty_line += 1

    cursor.close ()
    database.close ()

    full_now = str (datetime.now ()).replace (':', '-')
    dot_index = full_now.index ('.')
    now = full_now[:dot_index]
    
    wb.save ('invoices - {}.xlsx'.format (now))

if __name__ == '__main__':
    
    main ()



